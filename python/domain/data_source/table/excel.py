"""
    Obtain table data from an Excel file
"""
from __future__ import absolute_import
from __future__ import print_function


import os

import openpyxl
from openpyxl.utils import range_boundaries



class ExcelTableError( RuntimeError ):
    pass


class ExcelTable( object ):
    """
    Excel table data source
    """

    source_file = None
    work_book   = None
    sheet_name  = None
    cell_range  = None


    def __init__( self, filepath, sheet_name=None, cell_range=None ):
        """
        turn an Excel file into a table data source

        @param filepath: path to the Excel file
        @param sheet_name: name of the sheet within the Excel file that holds the table data
        @param cell_range: range of the cells within the sheet that hold the table data
        """
        if not os.path.isfile( filepath ):
            raise ExcelTableError( 'no file {}'.format(filepath) )

        self.source_file = filepath
        self.work_book   = openpyxl.load_workbook( filepath, read_only=True )
        self.set_sheet_as_source( sheet_name, cell_range )


    def set_sheet_as_source( self, sheet_name=None, cell_range=None ):
        """
        use sheet as source for table data
        
        @param sheet_name: name of the sheet
        @param cell_range: range of the cells within the sheet that hold the table data
        @note: defaults to all the data on the first sheet listed in the workbook
        """
        if sheet_name is None:
            sheet_name = self.work_book.sheetnames[0]
        elif sheet_name not in self.work_book.sheetnames:
            raise ExcelTableError( "no sheet named '{}' in {}".format( sheet_name, self.source_file) )

        self.sheet_name = sheet_name

        self.set_range_as_source( cell_range )


    def set_range_as_source( self, cell_range=None ):
        """
        use sheet as source for table data
        
        @param cell_range: range of the cells within the sheet that hold the table data
        @note: defaults to all the data on the chosen sheet
        """
        ws = self.work_sheet
        if cell_range is None:
            min_col, min_row, max_col, max_row = ws.min_column, ws.min_row, ws.max_column, ws.max_row
        elif isinstance( cell_range, basestring ):
            min_col, min_row, max_col, max_row = range_boundaries( cell_range )
        else:
            min_col, min_row, max_col, max_row = self.cell_range2components( cell_range )

        if not( ws.min_column <= min_col <= max_col <= ws.max_column ) or not( ws.min_row <= min_row <= max_row <= ws.max_row ):
            raise ExcelTableError( "cell range ({},{})-({},{}) not valid for sheet '{}' in {}".format( min_col, min_row, max_col, max_row, self.sheet_name, self.source_file) )

        self.cell_range = self.components2cell_range( min_col, min_row, max_col, max_row )


    # --- iteration interface ---

    def __iter__( self ):
        """
        produce rows of data
        """
        min_col, min_row, max_col, max_row = self.cell_range2components( self.cell_range )

        for row in self.work_sheet.iter_rows( min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row ):
            yield [ cell.value for cell in row ]


    # --- properties ---

    @property
    def work_sheet( self ):
        if self.sheet_name is None:
            raise ExcelTableError( 'no sheet name set' )
        elif self.sheet_name not in self.work_book.sheetnames:
            raise ExcelTableError( "no sheet named '{}' in {}".format( self.sheet_name, self.source_file) )

        return self.work_book[self.sheet_name]


    # --- utilities ---

    @staticmethod
    def components2cell_range( min_col, min_row, max_col, max_row ):
        cell_min   = ( min_col, min_row )
        cell_max   = ( max_col, max_row )
        cell_range = ( cell_min, cell_max )
        return cell_range
        

    @staticmethod
    def cell_range2components( cell_range ):
        cell_min, cell_max = cell_range
        min_col,  min_row  = cell_min
        max_col,  max_row  = cell_max
        return min_col, min_row, max_col, max_row
        
